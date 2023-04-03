from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import struct
import pyaudio
import pyttsx3
import speech_recognition as sr
import time

wb = Workbook()
wb = load_workbook("sample.xlsx")
ws = wb.active

a=""
while(True):
    break
    r = sr.Recognizer()
    with sr.Microphone() as source: 
        print("Listening ...")
        audio = r.listen(source)
        query = ''

        try:
            query = r.recognize_google(audio, language = 'en-IN', show_all=True)
            print(query['alternative'][0]['transcript'])
            a=  query['alternative'][0]['transcript'].lower()
            if a.__contains__("stop"):
                break
        # finally:
        #     pass
        except Exception as e:
            print("Speak loudly") 

import spacy
from spacy.tokens import DocBin
from tqdm import tqdm
from spacy.matcher import Matcher

operation = []
cell = []

nlp = spacy.load("content\model-best") 
transcriptions = "add a1 b1 store in c3"
doc = nlp(transcriptions)
print(transcriptions)
print(doc.ents)
matcher = Matcher(nlp.vocab)
pattern = [{"ENT_TYPE": "CELL"}]
#print(pattern[0]["ENT_TYPE"])
matcher.add("", [pattern])
matches = matcher(doc)
for match_id, start, end in matches:
    string_id = nlp.vocab.strings[match_id]  # Get string representation
    span = doc[start:end]  # The matched span
    if pattern[0]["ENT_TYPE"]=="CELL":
        cell.append(span.text.upper())
    if pattern[0]["ENT_TYPE"]=="OPERATION":
        operation.append(span.text)
    print( span.text)
if operation[0]=="add":
    ws[cell[2].value]=ws[cell[0].value]+ws[cell[1]]
    wb.save("sample.xlsx")


    
#spacy.displacy.serve(doc, style='ent')